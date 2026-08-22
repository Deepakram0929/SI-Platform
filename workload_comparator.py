import os
import tempfile
from io import BytesIO

import pandas as pd
import streamlit as st

from kubernetes import client, config
from kubernetes.client.exceptions import ApiException

from openpyxl.styles import PatternFill, Font, Alignment


ENVIRONMENTS = [
    "DEV",
    "UAT",
    "STAGING",
    "PREPROD",
    "PROD",
    "BLUE",
    "GREEN",
]


def create_kubernetes_clients(uploaded_file):

    temp_path = None

    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            suffix=".yaml",
            delete=False,
        ) as temp_file:

            temp_file.write(uploaded_file.getvalue())
            temp_path = temp_file.name

        config.load_kube_config(config_file=temp_path)

        return (
            client.CoreV1Api(),
            client.AppsV1Api(),
            client.BatchV1Api(),
        )

    finally:

        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


def get_deployments(apps_api):

    workloads = []

    try:
        response = apps_api.list_deployment_for_all_namespaces()

        for item in response.items:
            workloads.append(
                {
                    "namespace": item.metadata.namespace,
                    "type": "Deployment",
                    "name": item.metadata.name,
                }
            )

    except ApiException as e:
        st.warning(f"Unable to read Deployments: {e.reason}")

    return workloads


def get_statefulsets(apps_api):

    workloads = []

    try:
        response = apps_api.list_stateful_set_for_all_namespaces()

        for item in response.items:
            workloads.append(
                {
                    "namespace": item.metadata.namespace,
                    "type": "StatefulSet",
                    "name": item.metadata.name,
                }
            )

    except ApiException as e:
        st.warning(f"Unable to read StatefulSets: {e.reason}")

    return workloads


def get_daemonsets(apps_api):

    workloads = []

    try:
        response = apps_api.list_daemon_set_for_all_namespaces()

        for item in response.items:
            workloads.append(
                {
                    "namespace": item.metadata.namespace,
                    "type": "DaemonSet",
                    "name": item.metadata.name,
                }
            )

    except ApiException as e:
        st.warning(f"Unable to read DaemonSets: {e.reason}")

    return workloads


def get_cronjobs(batch_api):

    workloads = []

    try:
        response = batch_api.list_cron_job_for_all_namespaces()

        for item in response.items:
            workloads.append(
                {
                    "namespace": item.metadata.namespace,
                    "type": "CronJob",
                    "name": item.metadata.name,
                }
            )

        return workloads

    except ApiException as e:

        if e.status != 404:
            st.warning(f"Unable to read CronJobs: {e.reason}")

    try:
        batch_beta_api = client.BatchV1beta1Api()
        response = batch_beta_api.list_cron_job_for_all_namespaces()

        for item in response.items:
            workloads.append(
                {
                    "namespace": item.metadata.namespace,
                    "type": "CronJob",
                    "name": item.metadata.name,
                }
            )

    except Exception:
        pass

    return workloads


def get_all_workloads(uploaded_file):

    _, apps_api, batch_api = create_kubernetes_clients(
        uploaded_file
    )

    workloads = []

    workloads.extend(get_deployments(apps_api))
    workloads.extend(get_statefulsets(apps_api))
    workloads.extend(get_daemonsets(apps_api))
    workloads.extend(get_cronjobs(batch_api))

    df = pd.DataFrame(
        workloads,
        columns=["namespace", "type", "name"],
    )

    if not df.empty:
        df = (
            df.drop_duplicates()
            .sort_values(["namespace", "name", "type"])
            .reset_index(drop=True)
        )

    return df


def get_namespaces(env1_df, env2_df):

    namespaces = set()

    if not env1_df.empty:
        namespaces.update(
            env1_df["namespace"].dropna().tolist()
        )

    if not env2_df.empty:
        namespaces.update(
            env2_df["namespace"].dropna().tolist()
        )

    return sorted(namespaces)


def get_workload_names(df, namespace):

    if df.empty:
        return []

    namespace_df = df[
        df["namespace"] == namespace
    ]

    if namespace_df.empty:
        return []

    return sorted(
        namespace_df["name"]
        .dropna()
        .astype(str)
        .drop_duplicates()
        .tolist(),
        key=str.lower,
    )


def compare_namespace(namespace, env1_names, env2_names):

    rows = []

    common = sorted(
        set(env1_names) & set(env2_names),
        key=str.lower,
    )

    env1_only = sorted(
        set(env1_names) - set(env2_names),
        key=str.lower,
    )

    env2_only = sorted(
        set(env2_names) - set(env1_names),
        key=str.lower,
    )

    for name in common:
        rows.append(
            {
                "Namespace": namespace,
                "Environment 1": name,
                "Environment 2": name,
                "Status": "SAME",
            }
        )

    max_diff = max(
        len(env1_only),
        len(env2_only),
    )

    for index in range(max_diff):

        env1_name = (
            env1_only[index]
            if index < len(env1_only)
            else ""
        )

        env2_name = (
            env2_only[index]
            if index < len(env2_only)
            else ""
        )

        rows.append(
            {
                "Namespace": namespace,
                "Environment 1": env1_name,
                "Environment 2": env2_name,
                "Status": "DIFF",
            }
        )

    return rows


def create_comparison(
    env1_df,
    env2_df,
    selected_namespace,
):

    rows = []

    if selected_namespace != "All Namespaces":

        namespaces = [selected_namespace]

    else:

        namespaces = get_namespaces(
            env1_df,
            env2_df,
        )

    for namespace in namespaces:

        env1_names = get_workload_names(
            env1_df,
            namespace,
        )

        env2_names = get_workload_names(
            env2_df,
            namespace,
        )

        rows.extend(
            compare_namespace(
                namespace,
                env1_names,
                env2_names,
            )
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Namespace",
            "Environment 1",
            "Environment 2",
            "Status",
        ],
    )


def create_excel_report(
    comparison_df,
    env1_df,
    env2_df,
    env1,
    env2,
):

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        comparison_df.to_excel(
            writer,
            sheet_name="Comparison",
            index=False,
        )

        comparison_df[
            comparison_df["Status"] == "DIFF"
        ].to_excel(
            writer,
            sheet_name="Differences",
            index=False,
        )

        env1_df.to_excel(
            writer,
            sheet_name=f"{env1} Workloads"[:31],
            index=False,
        )

        env2_df.to_excel(
            writer,
            sheet_name=f"{env2} Workloads"[:31],
            index=False,
        )

        workbook = writer.book

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE",
        )

        red_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE",
        )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        worksheet = workbook["Comparison"]

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center"
            )

        status_column = None

        for cell in worksheet[1]:
            if cell.value == "Status":
                status_column = cell.column
                break

        if status_column:

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):

                status = worksheet.cell(
                    row=row_number,
                    column=status_column,
                ).value

                fill = (
                    green_fill
                    if status == "SAME"
                    else red_fill
                )

                for column_number in range(
                    1,
                    worksheet.max_column + 1,
                ):

                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).fill = fill

        for worksheet in workbook.worksheets:

            worksheet.freeze_panes = "A2"

            for column in worksheet.columns:

                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:

                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 4,
                    60,
                )

    output.seek(0)

    return output


def render_workload_comparator():

    st.markdown(
        '<div class="page-title">☸️ Workload Comparator</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="page-subtitle">'
        'Compare Kubernetes workload names between two environments.'
        '</div>',
        unsafe_allow_html=True,
    )

    upload_col1, upload_col2 = st.columns(2)

    with upload_col1:

        st.markdown("###  Environment 1")

        env1 = st.selectbox(
            "Select Environment 1",
            ENVIRONMENTS,
            index=1,
            key="workload_env1",
        )

        env1_file = st.file_uploader(
            f"Upload {env1} kubeconfig",
            type=["yaml", "yml", "conf"],
            key="workload_env1_file",
        )

    with upload_col2:

        st.markdown("###  Environment 2")

        env2 = st.selectbox(
            "Select Environment 2",
            ENVIRONMENTS,
            index=4,
            key="workload_env2",
        )

        env2_file = st.file_uploader(
            f"Upload {env2} kubeconfig",
            type=["yaml", "yml", "conf"],
            key="workload_env2_file",
        )

    if env1 == env2:
        st.warning(
            "Please select two different environments."
        )
        return

    if env1_file and env2_file:

        if st.button(
            f"🔗 Load {env1} & {env2} Workloads",
            use_container_width=True,
            key="load_workloads",
        ):

            try:

                with st.spinner(
                    f"Connecting to {env1}..."
                ):
                    st.session_state.workload_env1_df = (
                        get_all_workloads(env1_file)
                    )

                with st.spinner(
                    f"Connecting to {env2}..."
                ):
                    st.session_state.workload_env2_df = (
                        get_all_workloads(env2_file)
                    )

                st.session_state.workload_loaded_env1 = env1
                st.session_state.workload_loaded_env2 = env2
                st.session_state.workload_comparison = None

                st.success(
                    f"{env1} and {env2} workloads loaded successfully."
                )

            except Exception as e:

                st.error(
                    "Failed to connect to Kubernetes cluster."
                )
                st.exception(e)

    if (
        st.session_state.get("workload_env1_df") is not None
        and
        st.session_state.get("workload_env2_df") is not None
    ):

        env1_df = st.session_state.workload_env1_df
        env2_df = st.session_state.workload_env2_df

        namespaces = get_namespaces(
            env1_df,
            env2_df,
        )

        st.divider()

        selected_namespace = st.selectbox(
            "🔎 Select Namespace",
            ["All Namespaces"] + namespaces,
            key="workload_namespace",
        )

        if st.button(
            f"🔍 Compare {env1} vs {env2}",
            type="primary",
            use_container_width=True,
            key="compare_workloads",
        ):

            st.session_state.workload_comparison = (
                create_comparison(
                    env1_df,
                    env2_df,
                    selected_namespace,
                )
            )

    comparison_df = st.session_state.get(
        "workload_comparison"
    )

    if comparison_df is None:
        return

    st.divider()

    total = len(comparison_df)

    same = len(
        comparison_df[
            comparison_df["Status"] == "SAME"
        ]
    )

    diff = len(
        comparison_df[
            comparison_df["Status"] == "DIFF"
        ]
    )

    c1, c2, c3 = st.columns(3)

    c1.metric("Total", total)
    c2.metric("🟢 Same", same)
    c3.metric("🔴 Difference", diff)

    # Change column names to actual environments.
    display_df = comparison_df.rename(
        columns={
            "Environment 1": env1,
            "Environment 2": env2,
        }
    )

    def highlight_rows(row):

        if row["Status"] == "SAME":

            return [
                "background-color: #C6EFCE; "
                "color: #006100; "
                "font-weight: 600;"
            ] * len(row)

        return [
            "background-color: #FFC7CE; "
            "color: #9C0006; "
            "font-weight: 600;"
        ] * len(row)

    st.dataframe(
        display_df.style.apply(
            highlight_rows,
            axis=1,
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("### 🔴 Differences Only")

    differences_df = display_df[
        display_df["Status"] == "DIFF"
    ]

    if differences_df.empty:

        st.success(
            f"No differences found between {env1} and {env2}."
        )

    else:

        st.dataframe(
            differences_df.style.apply(
                lambda row: [
                    "background-color: #FFC7CE; "
                    "color: #9C0006; "
                    "font-weight: 600;"
                ] * len(row),
                axis=1,
            ),
            use_container_width=True,
            hide_index=True,
        )

    excel_data = create_excel_report(
        display_df,
        env1_df,
        env2_df,
        env1,
        env2,
    )

    st.download_button(
        "📥 Download Excel",
        data=excel_data,
        file_name=f"{env1}_{env2}_Workload_Comparison.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )