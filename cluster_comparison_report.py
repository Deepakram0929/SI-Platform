import base64
import re
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit as st
import yaml
import openpyxl

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from kubernetes import client
from kubernetes import config as kube_config


# ============================================================
# ENVIRONMENTS
# ============================================================

ENVIRONMENTS = [
    "UAT",
    "BLUE",
    "DEV",
    "Green",
    "STAGING",
    "Prod",
]


# ============================================================
# CREATE KUBERNETES CLIENTS
# ============================================================

def create_clients(kubeconfig_file):

    raw = kubeconfig_file.getvalue()

    if not raw:
        raise ValueError(
            "Kubeconfig file is empty."
        )

    kubeconfig_data = yaml.safe_load(
        raw.decode("utf-8")
    )

    api_client = (
        kube_config.new_client_from_config_dict(
            kubeconfig_data
        )
    )

    return {
        "core": client.CoreV1Api(
            api_client
        ),
        "apps": client.AppsV1Api(
            api_client
        ),
        "networking": client.NetworkingV1Api(
            api_client
        ),
    }


# ============================================================
# NAMESPACE LIST
# ============================================================

def get_namespaces(core_api):

    response = core_api.list_namespace(
        _request_timeout=30
    )

    return sorted(
        [
            item.metadata.name
            for item in response.items
            if item.metadata
            and item.metadata.name
        ],
        key=str.lower,
    )


# ============================================================
# WORKLOADS
# ============================================================

def collect_workloads(
    apps_api,
    namespace=None,
):

    rows = []

    # --------------------------------------------------------
    # DEPLOYMENTS
    # --------------------------------------------------------

    if namespace:

        deployments = (
            apps_api.list_namespaced_deployment(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        deployments = (
            apps_api.list_deployment_for_all_namespaces(
                _request_timeout=60,
            )
        )

    for item in deployments.items:

        ns = item.metadata.namespace
        workload = item.metadata.name

        for container in (
            item.spec.template.spec.containers
            or []
        ):

            rows.append(
                {
                    "Namespace": ns,
                    "Workload Type": "Deployment",
                    "Workload": workload,
                    "Container": container.name,
                    "Image": container.image or "",
                }
            )

    # --------------------------------------------------------
    # STATEFULSETS
    # --------------------------------------------------------

    if namespace:

        statefulsets = (
            apps_api.list_namespaced_stateful_set(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        statefulsets = (
            apps_api.list_stateful_set_for_all_namespaces(
                _request_timeout=60,
            )
        )

    for item in statefulsets.items:

        ns = item.metadata.namespace
        workload = item.metadata.name

        for container in (
            item.spec.template.spec.containers
            or []
        ):

            rows.append(
                {
                    "Namespace": ns,
                    "Workload Type": "StatefulSet",
                    "Workload": workload,
                    "Container": container.name,
                    "Image": container.image or "",
                }
            )

    # --------------------------------------------------------
    # DAEMONSETS
    # --------------------------------------------------------

    if namespace:

        daemonsets = (
            apps_api.list_namespaced_daemon_set(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        daemonsets = (
            apps_api.list_daemon_set_for_all_namespaces(
                _request_timeout=60,
            )
        )

    for item in daemonsets.items:

        ns = item.metadata.namespace
        workload = item.metadata.name

        for container in (
            item.spec.template.spec.containers
            or []
        ):

            rows.append(
                {
                    "Namespace": ns,
                    "Workload Type": "DaemonSet",
                    "Workload": workload,
                    "Container": container.name,
                    "Image": container.image or "",
                }
            )

    return rows


# ============================================================
# CONFIGMAP
# ============================================================

def collect_configmaps(
    core_api,
    namespace=None,
):
    """Collect ConfigMap values at key level for accurate comparison."""

    if namespace:
        response = core_api.list_namespaced_config_map(
            namespace=namespace,
            _request_timeout=45,
        )
    else:
        response = core_api.list_config_map_for_all_namespaces(
            _request_timeout=45,
        )

    rows = []

    for item in response.items:
        if not item.metadata:
            continue

        ns = item.metadata.namespace or ""
        name = item.metadata.name or ""

        # ConfigMap.data contains normal string values.
        for key, value in sorted((item.data or {}).items()):
            rows.append({
                "Namespace": ns,
                "ConfigMap": name,
                "Key": key,
                "Value": "" if value is None else str(value),
            })

        # Also include binaryData keys so they are not silently missed.
        for key, value in sorted((item.binary_data or {}).items()):
            rows.append({
                "Namespace": ns,
                "ConfigMap": name,
                "Key": key,
                "Value": "<binary data>",
            })

        # Keep empty ConfigMaps visible.
        if not item.data and not item.binary_data:
            rows.append({
                "Namespace": ns,
                "ConfigMap": name,
                "Key": "<NO DATA>",
                "Value": "",
            })

    return rows


# ============================================================
# SECRETS
# ============================================================

def _decode_secret_value(value):
    """Decode a Kubernetes Secret value when it is base64 encoded."""
    if value is None:
        return ""

    if isinstance(value, bytes):
        raw = value
    else:
        raw = str(value).encode("utf-8")

    try:
        decoded = base64.b64decode(raw, validate=True)
        return decoded.decode("utf-8", errors="replace")
    except Exception:
        # Some clients may already return decoded strings.
        return str(value)


def collect_secrets(
    core_api,
    namespace=None,
):
    """Collect Secret keys and decoded values at key level for comparison."""

    if namespace:
        response = core_api.list_namespaced_secret(
            namespace=namespace,
            _request_timeout=45,
        )
    else:
        response = core_api.list_secret_for_all_namespaces(
            _request_timeout=45,
        )

    rows = []

    for item in response.items:
        if not item.metadata:
            continue

        ns = item.metadata.namespace or ""
        name = item.metadata.name or ""
        data = item.data or {}

        for key, value in sorted(data.items()):
            rows.append({
                "Namespace": ns,
                "Secret": name,
                "Key": key,
                "Value": _decode_secret_value(value),
            })

        if not data:
            rows.append({
                "Namespace": ns,
                "Secret": name,
                "Key": "<NO DATA>",
                "Value": "",
            })

    return rows


# ============================================================
# INGRESS
# ============================================================

def collect_ingress(
    networking_api,
    namespace=None,
):

    if namespace:

        response = (
            networking_api.list_namespaced_ingress(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        response = (
            networking_api.list_ingress_for_all_namespaces(
                _request_timeout=60,
            )
        )

    rows = []

    for item in response.items:

        if not item.metadata:
            continue

        name = item.metadata.name
        ns = item.metadata.namespace

        rules = item.spec.rules or []

        if not rules:

            rows.append(
                {
                    "Namespace": ns,
                    "Ingress": name,
                    "Host": "",
                    "Path": "",
                    "Service": "",
                    "Port": "",
                }
            )

            continue

        for rule in rules:

            host = rule.host or ""

            if not rule.http:

                rows.append(
                    {
                        "Namespace": ns,
                        "Ingress": name,
                        "Host": host,
                        "Path": "",
                        "Service": "",
                        "Port": "",
                    }
                )

                continue

            for path in rule.http.paths:

                service = ""
                port = ""

                if path.backend.service:

                    service = (
                        path.backend.service.name
                        or ""
                    )

                    if path.backend.service.port:

                        if (
                            path.backend.service.port.number
                            is not None
                        ):

                            port = (
                                path.backend.service.port.number
                            )

                        else:

                            port = (
                                path.backend.service.port.name
                                or ""
                            )

                rows.append(
                    {
                        "Namespace": ns,
                        "Ingress": name,
                        "Host": host,
                        "Path": path.path or "",
                        "Service": service,
                        "Port": port,
                    }
                )

    return rows


# ============================================================
# PVC
# ============================================================

def collect_pvc(
    core_api,
    namespace=None,
):

    if namespace:

        response = (
            core_api.list_namespaced_persistent_volume_claim(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        response = (
            core_api.list_persistent_volume_claim_for_all_namespaces(
                _request_timeout=60,
            )
        )

    rows = []

    for item in response.items:

        if not item.metadata:
            continue

        storage = ""

        if item.spec.resources:

            requests = (
                item.spec.resources.requests
                or {}
            )

            storage = requests.get(
                "storage",
                "",
            )

        rows.append(
            {
                "Namespace":
                    item.metadata.namespace,
                "PVC":
                    item.metadata.name,
                "Status":
                    item.status.phase or "",
                "Storage Class":
                    item.spec.storage_class_name
                    or "",
                "Requested Storage":
                    storage,
                "Volume":
                    item.spec.volume_name
                    or "",
                "Access Modes":
                    ", ".join(
                        item.spec.access_modes
                        or []
                    ),
            }
        )

    return rows


# ============================================================
# VOLUMES
# ============================================================

def collect_volumes(
    apps_api,
    namespace=None,
):

    rows = []

    if namespace:

        deployments = (
            apps_api.list_namespaced_deployment(
                namespace=namespace,
                _request_timeout=60,
            )
        )

        statefulsets = (
            apps_api.list_namespaced_stateful_set(
                namespace=namespace,
                _request_timeout=60,
            )
        )

    else:

        deployments = (
            apps_api.list_deployment_for_all_namespaces(
                _request_timeout=60,
            )
        )

        statefulsets = (
            apps_api.list_stateful_set_for_all_namespaces(
                _request_timeout=60,
            )
        )

    workloads = []

    for item in deployments.items:

        workloads.append(
            (
                "Deployment",
                item,
            )
        )

    for item in statefulsets.items:

        workloads.append(
            (
                "StatefulSet",
                item,
            )
        )

    for workload_type, item in workloads:

        if not item.metadata:
            continue

        pod_spec = (
            item.spec.template.spec
        )

        volumes = (
            pod_spec.volumes or []
        )

        containers = (
            pod_spec.containers or []
        )

        for volume in volumes:

            volume_type = "Unknown"

            if volume.config_map:

                volume_type = "ConfigMap"

            elif volume.secret:

                volume_type = "Secret"

            elif volume.persistent_volume_claim:

                volume_type = "PVC"

            elif volume.empty_dir:

                volume_type = "EmptyDir"

            elif volume.host_path:

                volume_type = "HostPath"

            elif volume.projected:

                volume_type = "Projected"

            elif volume.nfs:

                volume_type = "NFS"

            mount_paths = []

            for container in containers:

                for mount in (
                    container.volume_mounts
                    or []
                ):

                    if mount.name == volume.name:

                        mount_paths.append(
                            f"{container.name}: "
                            f"{mount.mount_path}"
                        )

            rows.append(
                {
                    "Namespace":
                        item.metadata.namespace,
                    "Workload Type":
                        workload_type,
                    "Workload":
                        item.metadata.name,
                    "Volume":
                        volume.name,
                    "Volume Type":
                        volume_type,
                    "Mount Paths":
                        ", ".join(
                            mount_paths
                        ),
                }
            )

    return rows


# ============================================================
# FAST WORKLOAD + VOLUME COLLECTION
# ============================================================

def collect_workloads_and_volumes(
    apps_api,
    namespace=None,
    need_workloads=True,
    need_volumes=False,
):
    """
    Collect Deployments/StatefulSets/DaemonSets once and derive
    both Workloads and Volumes from the same API responses.

    This avoids the previous duplicate API calls where:
      - collect_workloads() fetched workloads
      - collect_volumes() fetched Deployments/StatefulSets again

    For ALL NAMESPACES this is significantly faster.
    """

    workload_rows = []
    volume_rows = []

    # --------------------------------------------------------
    # Fetch all workload types concurrently
    # --------------------------------------------------------

    def get_deployments():
        if namespace:
            return apps_api.list_namespaced_deployment(
                namespace=namespace,
                _request_timeout=45,
            )
        return apps_api.list_deployment_for_all_namespaces(
            _request_timeout=45,
        )

    def get_statefulsets():
        if namespace:
            return apps_api.list_namespaced_stateful_set(
                namespace=namespace,
                _request_timeout=45,
            )
        return apps_api.list_stateful_set_for_all_namespaces(
            _request_timeout=45,
        )

    def get_daemonsets():
        if namespace:
            return apps_api.list_namespaced_daemon_set(
                namespace=namespace,
                _request_timeout=45,
            )
        return apps_api.list_daemon_set_for_all_namespaces(
            _request_timeout=45,
        )

    with ThreadPoolExecutor(
        max_workers=3,
        thread_name_prefix="workload-api",
    ) as executor:

        future_deployments = executor.submit(
            get_deployments
        )
        future_statefulsets = executor.submit(
            get_statefulsets
        )
        future_daemonsets = executor.submit(
            get_daemonsets
        )

        deployments = future_deployments.result()
        statefulsets = future_statefulsets.result()
        daemonsets = future_daemonsets.result()

    workload_objects = [
        ("Deployment", item)
        for item in deployments.items
    ]

    workload_objects.extend(
        [
            ("StatefulSet", item)
            for item in statefulsets.items
        ]
    )

    workload_objects.extend(
        [
            ("DaemonSet", item)
            for item in daemonsets.items
        ]
    )

    # --------------------------------------------------------
    # Build workload rows
    # --------------------------------------------------------

    if need_workloads:

        for workload_type, item in workload_objects:

            if not item.metadata:
                continue

            ns = item.metadata.namespace or ""
            workload = item.metadata.name or ""

            pod_spec = (
                item.spec.template.spec
                if item.spec
                and item.spec.template
                and item.spec.template.spec
                else None
            )

            containers = (
                pod_spec.containers
                if pod_spec
                else []
            ) or []

            for container in containers:

                workload_rows.append(
                    {
                        "Namespace": ns,
                        "Workload Type": workload_type,
                        "Workload": workload,
                        "Container": container.name or "",
                        "Image": container.image or "",
                    }
                )

    # --------------------------------------------------------
    # Build volume rows from the SAME workload objects
    # --------------------------------------------------------

    if need_volumes:

        for workload_type, item in workload_objects:

            if not item.metadata:
                continue

            pod_spec = (
                item.spec.template.spec
                if item.spec
                and item.spec.template
                and item.spec.template.spec
                else None
            )

            if not pod_spec:
                continue

            volumes = (
                pod_spec.volumes or []
            )

            containers = (
                pod_spec.containers or []
            )

            for volume in volumes:

                volume_type = "Unknown"

                if volume.config_map:
                    volume_type = "ConfigMap"

                elif volume.secret:
                    volume_type = "Secret"

                elif volume.persistent_volume_claim:
                    volume_type = "PVC"

                elif volume.empty_dir:
                    volume_type = "EmptyDir"

                elif volume.host_path:
                    volume_type = "HostPath"

                elif volume.projected:
                    volume_type = "Projected"

                elif volume.nfs:
                    volume_type = "NFS"

                mount_paths = []

                for container in containers:

                    for mount in (
                        container.volume_mounts or []
                    ):

                        if mount.name == volume.name:

                            mount_paths.append(
                                f"{container.name}: "
                                f"{mount.mount_path}"
                            )

                volume_rows.append(
                    {
                        "Namespace":
                            item.metadata.namespace or "",
                        "Workload Type":
                            workload_type,
                        "Workload":
                            item.metadata.name or "",
                        "Volume":
                            volume.name or "",
                        "Volume Type":
                            volume_type,
                        "Mount Paths":
                            ", ".join(
                                mount_paths
                            ),
                    }
                )

    return workload_rows, volume_rows


# ============================================================
# GENERIC COMPARISON
# ============================================================

def compare_rows(
    source_rows,
    destination_rows,
    key_columns,
):

    source_map = {
        tuple(
            row.get(column, "")
            for column in key_columns
        ): row
        for row in source_rows
    }

    destination_map = {
        tuple(
            row.get(column, "")
            for column in key_columns
        ): row
        for row in destination_rows
    }

    all_keys = sorted(
        set(source_map)
        |
        set(destination_map),
        key=str,
    )

    result = []

    for key in all_keys:

        source = source_map.get(
            key
        )

        destination = (
            destination_map.get(
                key
            )
        )

        if source and destination:

            source_values = {
                k: v
                for k, v in source.items()
                if k not in key_columns
            }

            destination_values = {
                k: v
                for k, v in destination.items()
                if k not in key_columns
            }

            status = (
                "SAME"
                if source_values
                == destination_values
                else "DIFF"
            )

        elif source:

            status = "MISSING"

        else:

            status = "DESTINATION ONLY"

        row = {}

        for column in key_columns:

            row[column] = key[
                key_columns.index(column)
            ]

        if source:

            for column, value in source.items():

                if column not in key_columns:

                    row[
                        f"Source {column}"
                    ] = value

        if destination:

            for column, value in destination.items():

                if column not in key_columns:

                    row[
                        f"Destination {column}"
                    ] = value

        row["Status"] = status

        result.append(row)

    return result


# ============================================================
# EXCEL
# ============================================================

def _excel_safe_value(value):
    """Remove XML control characters that Excel/openpyxl cannot store."""
    if value is None:
        return ""
    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    if isinstance(value, (int, float, bool, datetime)):
        return value
    return str(value)


def _excel_safe_row(values):
    return [_excel_safe_value(value) for value in values]


def _excel_styles():
    thin = Side(style="thin", color="7F7F7F")
    return {
        "header_fill": PatternFill("solid", fgColor="4F81BD"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "same_fill": PatternFill("solid", fgColor="C6E0B4"),
        "diff_fill": PatternFill("solid", fgColor="F4B183"),
        "missing_fill": PatternFill("solid", fgColor="FFE699"),
        "destination_fill": PatternFill("solid", fgColor="BDD7EE"),
        "white_fill": PatternFill("solid", fgColor="FFFFFF"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "thin": thin,
    }


def _apply_status(cell, status, styles):
    status = str(status or "").strip().upper()

    if status == "SAME":
        cell.fill = styles["same_fill"]
        cell.font = Font(bold=True, color="548235")
    elif status in {"DIFF", "CHANGED"}:
        cell.fill = styles["diff_fill"]
        cell.font = Font(bold=True, color="C65911")
    elif status == "MISSING":
        cell.fill = styles["missing_fill"]
        cell.font = Font(bold=True, color="9C6500")
    elif status == "DESTINATION ONLY":
        cell.fill = styles["destination_fill"]
        cell.font = Font(bold=True, color="2F75B5")


def _finish_sheet(ws, headers, rows, status_column=None, freeze="A2"):
    styles = _excel_styles()
    ws.sheet_view.showGridLines = False

    # Header
    for cell in ws[1]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Data
    for row_index in range(2, ws.max_row + 1):
        for cell in ws[row_index]:
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFF")

        if status_column:
            _apply_status(ws.cell(row_index, status_column), ws.cell(row_index, status_column).value, styles)

    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"

    # Widths. Long values are kept readable but do not make the workbook enormous.
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        h = str(header).lower()
        width = max(12, min(32, len(str(header)) + 4))

        if "namespace" in h:
            width = 24
        elif h in {"configmap", "secret", "ingress", "pvc", "workload"}:
            width = 34
        elif h == "key":
            width = 38
        elif h == "value" or "image" in h:
            width = 58
        elif "container" in h:
            width = 30
        elif "status" in h:
            width = 20
        elif "host" in h:
            width = 38
        elif "path" in h:
            width = 34
        elif "service" in h:
            width = 32
        elif "mount" in h:
            width = 42

        ws.column_dimensions[letter].width = width


def add_sheet(workbook, name, rows):
    """Create a normal comparison table with borders and status colors."""
    ws = workbook.create_sheet(name)
    styles = _excel_styles()

    if not rows:
        ws.append(_excel_safe_row(["No data found"]))
        c = ws["A1"]
        c.fill = styles["header_fill"]
        c.font = styles["header_font"]
        c.border = styles["border"]
        ws.column_dimensions["A"].width = 30
        ws.sheet_view.showGridLines = False
        return ws

    headers = list(rows[0].keys())
    ws.append(_excel_safe_row(headers))
    for row in rows:
        ws.append(_excel_safe_row([row.get(h, "") for h in headers]))

    status_column = next(
        (i for i, h in enumerate(headers, start=1) if str(h).lower() == "status"),
        None,
    )
    _finish_sheet(ws, headers, rows, status_column)
    return ws


def add_key_value_comparison_sheet(
    workbook,
    name,
    source_rows,
    destination_rows,
    source_env,
    destination_env,
    resource_column,
):
    """
    ConfigMap/Secret format:

    Namespace | ConfigMap/Secret | source env: BLUE | value |
    Destination env: Green | value | Status

    One row is produced for every key, so individual values are compared
    instead of comparing a giant comma-separated key list.
    """
    ws = workbook.create_sheet(name)
    styles = _excel_styles()

    source_map = {}
    destination_map = {}

    for row in source_rows:
        key = (row.get("Namespace", ""), row.get(resource_column, ""), row.get("Key", ""))
        source_map[key] = row.get("Value", "")

    for row in destination_rows:
        key = (row.get("Namespace", ""), row.get(resource_column, ""), row.get("Key", ""))
        destination_map[key] = row.get("Value", "")

    all_keys = sorted(set(source_map) | set(destination_map), key=lambda x: tuple(str(v).lower() for v in x))

    headers = [
        "Namespace",
        resource_column,
        f"source env: {source_env}",
        "value",
        f"Destination env: {destination_env}",
        "value ",
        "Status",
    ]
    ws.append(_excel_safe_row(headers))

    for namespace, resource, key in all_keys:
        source_exists = (namespace, resource, key) in source_map
        destination_exists = (namespace, resource, key) in destination_map
        source_value = source_map.get((namespace, resource, key), "missing")
        destination_value = destination_map.get((namespace, resource, key), "missing")

        if source_exists and destination_exists:
            status = "SAME" if source_value == destination_value else "DIFF"
        elif source_exists:
            status = "MISSING"
        else:
            status = "DESTINATION ONLY"

        ws.append(_excel_safe_row([
            namespace, resource, key, source_value,
            key if destination_exists else "missing", destination_value, status,
        ]))

    _finish_sheet(ws, headers, all_keys, status_column=7)

    # Highlight the actual differing source/destination values.
    for row_index in range(2, ws.max_row + 1):
        status = str(ws.cell(row_index, 7).value or "").upper()
        if status == "DIFF":
            ws.cell(row_index, 4).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row_index, 6).fill = PatternFill("solid", fgColor="F4CCCC")
        elif status == "MISSING":
            ws.cell(row_index, 4).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row_index, 6).fill = styles["missing_fill"]
        elif status == "DESTINATION ONLY":
            ws.cell(row_index, 4).fill = styles["destination_fill"]
            ws.cell(row_index, 6).fill = PatternFill("solid", fgColor="D9EAF7")

    return ws


def add_ingress_comparison_sheet(
    workbook,
    source_rows,
    destination_rows,
    source_env,
    destination_env,
):
    """Create the two-block Ingress sheet shown in the requested format."""
    ws = workbook.create_sheet("Ingress")
    styles = _excel_styles()

    headers = [
        "Namespace", "Ingress", "Host", "Path", "Service", "Destination Port", "Status"
    ]

    # Match rows by namespace/ingress/path/service/port-independent key.
    def key(row):
        return (
            row.get("Namespace", ""),
            row.get("Ingress", ""),
            row.get("Path", ""),
            row.get("Service", ""),
            str(row.get("Port", "")),
        )

    source_map = {key(r): r for r in source_rows}
    destination_map = {key(r): r for r in destination_rows}

    all_keys = sorted(set(source_map) | set(destination_map), key=lambda x: tuple(str(v).lower() for v in x))

    # Source block
    ws.append(_excel_safe_row([f"Source ({source_env})"]))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    source_title = ws.cell(1, 1)
    source_title.fill = styles["header_fill"]
    source_title.font = styles["header_font"]
    source_title.alignment = Alignment(horizontal="left")
    source_title.border = styles["border"]

    ws.append(_excel_safe_row(headers[:-1]))
    for row in source_rows:
        ws.append(_excel_safe_row([
            row.get("Namespace", ""), row.get("Ingress", ""),
            row.get("Host", ""), row.get("Path", ""),
            row.get("Service", ""), row.get("Port", ""),
        ]))

    source_end = ws.max_row

    # Blank row + destination block
    ws.append([])
    dest_title_row = ws.max_row + 1
    ws.append(_excel_safe_row([f"Destination ({destination_env})"]))
    ws.merge_cells(start_row=dest_title_row, start_column=1, end_row=dest_title_row, end_column=len(headers))
    title = ws.cell(dest_title_row, 1)
    title.fill = styles["header_fill"]
    title.font = styles["header_font"]
    title.alignment = Alignment(horizontal="left")
    title.border = styles["border"]

    ws.append(_excel_safe_row(headers))
    dest_header_row = ws.max_row

    for k in all_keys:
        row = destination_map.get(k)
        source = source_map.get(k)

        if row is None:
            continue

        status = "SAME" if source and all(
            str(source.get(field, "")) == str(row.get(field, ""))
            for field in ("Host", "Path", "Service", "Port")
        ) else "DIFF" if source else "DESTINATION ONLY"

        ws.append(_excel_safe_row([
            row.get("Namespace", ""), row.get("Ingress", ""),
            row.get("Host", ""), row.get("Path", ""),
            row.get("Service", ""), row.get("Port", ""), status,
        ]))

        current = ws.max_row
        if status == "DIFF" and source:
            # Mark destination fields that differ from source.
            mapping = {2: "Ingress", 3: "Host", 4: "Path", 5: "Service", 6: "Port"}
            for col, field in mapping.items():
                if str(source.get(field, "")) != str(row.get(field, "")):
                    ws.cell(current, col).fill = PatternFill("solid", fgColor="F4CCCC")

    # Source and destination table formatting.
    for row_index in range(2, source_end + 1):
        for cell in ws[row_index]:
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(dest_header_row, ws.max_row + 1):
        for cell in ws[row_index]:
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[2]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = styles["border"]

    for cell in ws[dest_header_row]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = styles["border"]

    # Status colors in destination block.
    for row_index in range(dest_header_row + 1, ws.max_row + 1):
        _apply_status(ws.cell(row_index, 7), ws.cell(row_index, 7).value, styles)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    widths = [24, 30, 40, 34, 34, 20, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    return ws


def _add_namespace_dropdown(workbook, namespaces, summary_cell="B5"):
    """Add an Excel namespace dropdown to the Summary sheet.

    The dropdown is populated from a hidden Lists sheet.  It contains only
    namespaces included in this report.  A first option, ALL SELECTED, is
    provided when more than one namespace is present.
    """
    if "Summary" not in workbook.sheetnames:
        return

    summary = workbook["Summary"]
    values = [str(n) for n in (namespaces or []) if str(n).strip()]
    values = sorted(set(values), key=str.lower)

    if not values:
        return

    # Reuse/create hidden worksheet containing the validation values.
    if "Lists" in workbook.sheetnames:
        lists_ws = workbook["Lists"]
        lists_ws.delete_rows(1, lists_ws.max_row)
    else:
        lists_ws = workbook.create_sheet("Lists")

    lists_ws.sheet_state = "hidden"
    lists_ws["A1"] = "ALL SELECTED"
    for idx, namespace in enumerate(values, start=2):
        lists_ws.cell(idx, 1, namespace)

    last_row = len(values) + 1
    formula = f"='Lists'!$A$1:$A${last_row}"

    # Remove previous validations from Summary to avoid duplicate controls.
    summary.data_validations.dataValidation = []

    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
    )
    validation.error = "Select a namespace from the list."
    validation.errorTitle = "Invalid namespace"
    validation.prompt = "Choose a namespace to review in this report."
    validation.promptTitle = "Namespace"

    summary.add_data_validation(validation)
    validation.add(summary[summary_cell])

    # Add a clear label beside the dropdown.
    summary["A5"] = "Namespace Filter"
    summary[summary_cell] = "ALL SELECTED" if len(values) > 1 else values[0]
    summary["A5"].font = Font(bold=True)
    summary[summary_cell].font = Font(bold=True, color="1F4E78")
    summary[summary_cell].fill = PatternFill("solid", fgColor="EAF2F8")
    summary[summary_cell].border = _excel_styles()["border"]

    # Add a short instruction.
    summary["A7"] = "Use the dropdown to select a namespace. Resource sheets also have Excel filters on the Namespace column."
    summary.merge_cells("A7:F7")
    summary["A7"].font = Font(italic=True, color="666666")
    summary["A7"].alignment = Alignment(wrap_text=True)
    summary.row_dimensions[7].height = 30

    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 55


def _add_namespace_filter_buttons(workbook, namespaces):
    """Make Namespace columns filterable in every detailed sheet."""
    # openpyxl's AutoFilter is already enabled by _finish_sheet/add_sheet.
    # This helper only documents the available namespaces in the workbook.
    return


def generate_excel(
    source_env,
    destination_env,
    scope,
    namespace,
    source,
    destination,
):

    workbook = openpyxl.Workbook()

    workbook.remove(
        workbook.active
    )

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    summary = [
        {
            "Property":
                "Source Environment",
            "Value":
                source_env,
        },
        {
            "Property":
                "Destination Environment",
            "Value":
                destination_env,
        },
        {
            "Property":
                "Scope",
            "Value":
                scope,
        },
        {
            "Property":
                "Namespace(s)",
            "Value":
                (
                    ", ".join(namespace)
                    if isinstance(namespace, (list, tuple, set)) and namespace
                    else (namespace or "ALL NAMESPACES")
                ),
        },
        {
            "Property":
                "Generated",
            "Value":
                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
        },
    ]

    add_sheet(
        workbook,
        "Summary",
        summary,
    )

    # Excel namespace dropdown. Only namespaces included in this report
    # are offered in the dropdown.
    report_namespaces = []
    if isinstance(namespace, (list, tuple, set)):
        report_namespaces = list(namespace)
    elif namespace:
        report_namespaces = [namespace]
    else:
        # For ALL NAMESPACES, derive the complete namespace list from the
        # collected workload/configmap/secret data.
        for resource_name in (
            "workloads", "configmaps", "secrets",
            "ingress", "volumes", "pvc"
        ):
            for row in source.get(resource_name, []) or []:
                ns = row.get("Namespace")
                if ns:
                    report_namespaces.append(ns)
            for row in destination.get(resource_name, []) or []:
                ns = row.get("Namespace")
                if ns:
                    report_namespaces.append(ns)
        report_namespaces = sorted(set(report_namespaces), key=str.lower)

    _add_namespace_dropdown(
        workbook,
        report_namespaces,
        summary_cell="B5",
    )

    # --------------------------------------------------------
    # WORKLOADS
    # --------------------------------------------------------

    workload_rows = compare_rows(
        source["workloads"],
        destination["workloads"],
        [
            "Namespace",
            "Workload Type",
            "Workload",
            "Container",
        ],
    )

    add_sheet(
        workbook,
        "Workloads",
        workload_rows,
    )

    # --------------------------------------------------------
    # IMAGES
    # --------------------------------------------------------

    image_rows = []

    for row in workload_rows:

        image_rows.append(
            {
                "Namespace":
                    row.get(
                        "Namespace",
                        "",
                    ),
                "Workload Type":
                    row.get(
                        "Workload Type",
                        "",
                    ),
                "Workload":
                    row.get(
                        "Workload",
                        "",
                    ),
                "Container":
                    row.get(
                        "Container",
                        "",
                    ),
                "Source Image":
                    row.get(
                        "Source Image",
                        "",
                    ),
                "Destination Image":
                    row.get(
                        "Destination Image",
                        "",
                    ),
                "Status":
                    row.get(
                        "Status",
                        "",
                    ),
            }
        )

    add_sheet(
        workbook,
        "Images",
        image_rows,
    )

    # --------------------------------------------------------
    # CONFIGMAP - KEY LEVEL
    # --------------------------------------------------------

    add_key_value_comparison_sheet(
        workbook,
        "ConfigMaps",
        source["configmaps"],
        destination["configmaps"],
        source_env,
        destination_env,
        "ConfigMap",
    )

    # --------------------------------------------------------
    # SECRETS - KEY LEVEL
    # --------------------------------------------------------

    add_key_value_comparison_sheet(
        workbook,
        "Secrets",
        source["secrets"],
        destination["secrets"],
        source_env,
        destination_env,
        "Secret",
    )

    # --------------------------------------------------------
    # INGRESS - SOURCE / DESTINATION BLOCKS
    # --------------------------------------------------------

    add_ingress_comparison_sheet(
        workbook,
        source["ingress"],
        destination["ingress"],
        source_env,
        destination_env,
    )

    # --------------------------------------------------------
    # VOLUMES
    # --------------------------------------------------------

    volume_rows = compare_rows(
        source["volumes"],
        destination["volumes"],
        [
            "Namespace",
            "Workload Type",
            "Workload",
            "Volume",
        ],
    )

    add_sheet(
        workbook,
        "Volumes",
        volume_rows,
    )

    # --------------------------------------------------------
    # PVC
    # --------------------------------------------------------

    pvc_rows = compare_rows(
        source["pvcs"],
        destination["pvcs"],
        [
            "Namespace",
            "PVC",
        ],
    )

    add_sheet(
        workbook,
        "PVC",
        pvc_rows,
    )

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output


# ============================================================
# COLLECT REPORT DATA - PARALLEL / FAST
# ============================================================

def _collect_report_single_namespace(
    clients,
    namespace=None,
    selected_resources=None,
):
    """
    FAST cluster scan.

    Optimizations:
      1. Independent Kubernetes APIs run concurrently.
      2. Deployments/StatefulSets/DaemonSets are fetched once.
      3. Images are derived from workload data.
      4. Volumes are derived from the same workload data.
      5. ALL NAMESPACES uses cluster-wide list APIs directly.
    """

    selected_resources = set(
        selected_resources
        or [
            "Workloads",
            "Images",
            "ConfigMaps",
            "Secrets",
            "Ingress",
            "Volumes",
            "PVC",
        ]
    )

    data = {
        "workloads": [],
        "configmaps": [],
        "secrets": [],
        "ingress": [],
        "volumes": [],
        "pvcs": [],
    }

    # --------------------------------------------------------
    # Workloads, Images and Volumes share the same API calls.
    # --------------------------------------------------------

    need_workloads = (
        "Workloads" in selected_resources
        or "Images" in selected_resources
    )

    need_volumes = (
        "Volumes" in selected_resources
    )

    jobs = {}

    if need_workloads or need_volumes:

        jobs["workloads_and_volumes"] = (
            lambda: collect_workloads_and_volumes(
                clients["apps"],
                namespace,
                need_workloads=(
                    need_workloads
                ),
                need_volumes=(
                    need_volumes
                ),
            )
        )

    if "ConfigMaps" in selected_resources:

        jobs["configmaps"] = (
            lambda: collect_configmaps(
                clients["core"],
                namespace,
            )
        )

    if "Secrets" in selected_resources:

        jobs["secrets"] = (
            lambda: collect_secrets(
                clients["core"],
                namespace,
            )
        )

    if "Ingress" in selected_resources:

        jobs["ingress"] = (
            lambda: collect_ingress(
                clients["networking"],
                namespace,
            )
        )

    if "PVC" in selected_resources:

        jobs["pvcs"] = (
            lambda: collect_pvc(
                clients["core"],
                namespace,
            )
        )

    # --------------------------------------------------------
    # Run ALL independent jobs concurrently.
    # --------------------------------------------------------

    if jobs:

        max_workers = min(
            8,
            len(jobs),
        )

        with ThreadPoolExecutor(
            max_workers=max_workers,
            thread_name_prefix="k8s-report",
        ) as executor:

            future_map = {
                executor.submit(job): name
                for name, job in jobs.items()
            }

            for future in as_completed(
                future_map
            ):

                name = future_map[
                    future
                ]

                result = future.result()

                if name == "workloads_and_volumes":

                    (
                        workload_rows,
                        volume_rows,
                    ) = result

                    if need_workloads:

                        data[
                            "workloads"
                        ] = workload_rows

                    if need_volumes:

                        data[
                            "volumes"
                        ] = volume_rows

                else:

                    data[name] = result

    return data


def _merge_report_data(target, source):
    """Merge per-namespace report data without duplicating rows."""
    for key in target:
        target[key].extend(source.get(key, []))


def _deduplicate_rows(rows):
    """Remove duplicate report rows while preserving order."""
    seen = set()
    result = []

    for row in rows:
        marker = tuple(
            (key, str(row.get(key, "")))
            for key in sorted(row.keys())
        )
        if marker not in seen:
            seen.add(marker)
            result.append(row)

    return result


def collect_report(
    clients,
    namespace=None,
    selected_resources=None,
    namespaces=None,
):
    """
    Collect report data for one or multiple namespaces.

    namespace is kept for backward compatibility.
    For multiple selected namespaces, each namespace is queried in
    parallel so the application does not have to scan the complete
    cluster.
    """

    if namespaces is None:
        if isinstance(namespace, (list, tuple, set)):
            namespaces = list(namespace)
            namespace = None
        elif namespace:
            namespaces = [namespace]
        else:
            namespaces = None

    # ALL namespaces: use the existing optimized cluster-wide APIs.
    if namespaces is None:
        return _collect_report_single_namespace(
            clients,
            None,
            selected_resources,
        )

    selected_namespaces = sorted(
        {str(ns) for ns in namespaces if ns},
        key=str.lower,
    )

    if not selected_namespaces:
        return {
            "workloads": [],
            "configmaps": [],
            "secrets": [],
            "ingress": [],
            "volumes": [],
            "pvcs": [],
        }

    merged = {
        "workloads": [],
        "configmaps": [],
        "secrets": [],
        "ingress": [],
        "volumes": [],
        "pvcs": [],
    }

    # One namespace = one set of namespace-scoped API calls.
    # Multiple namespaces run concurrently.
    workers = min(12, max(1, len(selected_namespaces)))

    with ThreadPoolExecutor(
        max_workers=workers,
        thread_name_prefix="namespace-report",
    ) as executor:

        futures = {
            executor.submit(
                _collect_report_single_namespace,
                clients,
                ns,
                selected_resources,
            ): ns
            for ns in selected_namespaces
        }

        for future in as_completed(futures):
            ns = futures[future]
            try:
                _merge_report_data(merged, future.result())
            except Exception as exc:
                raise RuntimeError(
                    f"Failed to collect namespace '{ns}': {exc}"
                ) from exc

    for key in merged:
        merged[key] = _deduplicate_rows(merged[key])

    return merged


def collect_both_clusters(
    source_clients,
    destination_clients,
    namespace=None,
    selected_resources=None,
    namespaces=None,
):
    """
    FAST source + destination scan for one, multiple, or all namespaces.

    Source and destination clusters are scanned simultaneously.
    When multiple namespaces are selected, namespace scans are also
    parallelized inside each cluster.
    """

    if namespaces is None:
        if isinstance(namespace, (list, tuple, set)):
            namespaces = list(namespace)
            namespace = None
        elif namespace:
            namespaces = [namespace]
        else:
            namespaces = None

    with ThreadPoolExecutor(
        max_workers=2,
        thread_name_prefix="cluster-report",
    ) as executor:

        source_future = executor.submit(
            collect_report,
            source_clients,
            None if namespaces is not None else namespace,
            selected_resources,
            namespaces,
        )

        destination_future = executor.submit(
            collect_report,
            destination_clients,
            None if namespaces is not None else namespace,
            selected_resources,
            namespaces,
        )

        source_data = source_future.result()
        destination_data = destination_future.result()

    return source_data, destination_data


# ============================================================
# MAIN REPORT PAGE
# ============================================================

def render_cluster_comparison_report():

    st.title(
        "📊 Cluster Comparison Report"
    )

    st.caption(
        "Independent cluster comparison and "
        "Excel report generation."
    )

    # ========================================================
    # ENVIRONMENT
    # ========================================================

    col1, col2 = st.columns(2)

    with col1:

        source_env = st.selectbox(
            "Source",
            ENVIRONMENTS,
            index=ENVIRONMENTS.index(
                "BLUE"
            ),
            key="report_source_env",
        )

    with col2:

        destination_options = [
            env
            for env in ENVIRONMENTS
            if env != source_env
        ]

        default_index = (
            destination_options.index(
                "Green"
            )
            if "Green"
            in destination_options
            else 0
        )

        destination_env = st.selectbox(
            "Destination",
            destination_options,
            index=default_index,
            key="report_destination_env",
        )

    # ========================================================
    # KUBECONFIG
    # ========================================================

    col1, col2 = st.columns(2)

    with col1:

        source_file = st.file_uploader(
            f"{source_env} Kubeconfig",
            type=[
                "yaml",
                "yml",
                "conf",
            ],
            key="report_source_file",
        )

    with col2:

        destination_file = st.file_uploader(
            f"{destination_env} Kubeconfig",
            type=[
                "yaml",
                "yml",
                "conf",
            ],
            key="report_destination_file",
        )

    # ========================================================
    # CONNECT
    # ========================================================

    if st.button(
        "🔗 Connect Both Clusters",
        type="primary",
        use_container_width=True,
        key="report_connect",
    ):

        if (
            source_file is None
            or destination_file is None
        ):

            st.error(
                "Please upload both kubeconfig files."
            )

        else:

            try:

                with st.spinner(
                    "Connecting to clusters..."
                ):

                    source_clients = (
                        create_clients(
                            source_file
                        )
                    )

                    destination_clients = (
                        create_clients(
                            destination_file
                        )
                    )

                    # Load namespace lists concurrently.
                    with ThreadPoolExecutor(
                        max_workers=2,
                        thread_name_prefix="namespace-load",
                    ) as executor:

                        source_namespace_future = (
                            executor.submit(
                                get_namespaces,
                                source_clients["core"],
                            )
                        )

                        destination_namespace_future = (
                            executor.submit(
                                get_namespaces,
                                destination_clients["core"],
                            )
                        )

                        source_namespaces = (
                            source_namespace_future.result()
                        )

                        destination_namespaces = (
                            destination_namespace_future.result()
                        )

                    common_namespaces = sorted(
                        set(
                            source_namespaces
                        )
                        &
                        set(
                            destination_namespaces
                        ),
                        key=str.lower,
                    )

                    st.session_state[
                        "report_source_clients"
                    ] = source_clients

                    st.session_state[
                        "report_destination_clients"
                    ] = destination_clients

                    st.session_state[
                        "report_namespaces"
                    ] = common_namespaces

                st.success(
                    "Both clusters connected."
                )

            except Exception as exc:

                st.error(
                    "Cluster connection failed."
                )

                st.exception(exc)

    # ========================================================
    # CHECK CONNECTION
    # ========================================================

    source_clients = st.session_state.get(
        "report_source_clients"
    )

    destination_clients = st.session_state.get(
        "report_destination_clients"
    )

    namespaces = st.session_state.get(
        "report_namespaces",
        [],
    )

    if (
        source_clients is None
        or destination_clients is None
    ):

        return

    # ========================================================
    # SCOPE
    # ========================================================

    st.markdown(
        "### 📁 Comparison Scope"
    )

    scope = st.radio(
        "Select scope",
        [
            "Selected Namespace",
            "All Namespaces",
        ],
        horizontal=True,
        key="report_scope",
    )

    selected_namespaces = []

    if scope == "Selected Namespace":

        if not namespaces:

            st.warning(
                "No common namespaces found."
            )

            return

        # MULTI-NAMESPACE SELECTION
        # Users can select any number of namespaces. Only these namespaces
        # are sent to the report collector; the rest of the cluster is not scanned.
        #
        # IMPORTANT: Do not modify st.session_state["report_namespaces_selected"]
        # directly after st.multiselect() has been created. Streamlit protects
        # widget-owned session-state keys during the current script run.
        # The callbacks below run BEFORE the next script run, so Select All/Clear
        # can safely update the widget value without StreamlitAPIException.

        def _select_all_report_namespaces():
            st.session_state["report_namespaces_selected"] = list(namespaces)

        def _clear_report_namespaces():
            st.session_state["report_namespaces_selected"] = []

        # Initialize the widget state BEFORE the widget is instantiated.
        if "report_namespaces_selected" not in st.session_state:
            st.session_state["report_namespaces_selected"] = []
        else:
            # Keep only namespaces that are currently available in both clusters.
            # This prevents stale selections after reconnecting clusters.
            st.session_state["report_namespaces_selected"] = [
                ns
                for ns in st.session_state["report_namespaces_selected"]
                if ns in namespaces
            ]

        selected_namespaces = st.multiselect(
            "Select Namespace(s)",
            options=namespaces,
            key="report_namespaces_selected",
            placeholder="Select one or more namespaces...",
            help=(
                "Select multiple specific namespaces. "
                "Only the selected namespaces will be scanned and included in the Excel report."
            ),
        )

        # Convenience buttons use callbacks. They must NOT assign the widget key
        # in the normal body of the script after multiselect() is instantiated.
        c1, c2 = st.columns([1, 1])
        with c1:
            st.button(
                "Select All Listed Namespaces",
                key="report_select_all_namespaces",
                use_container_width=True,
                on_click=_select_all_report_namespaces,
            )
        with c2:
            st.button(
                "Clear Namespace Selection",
                key="report_clear_namespaces",
                use_container_width=True,
                on_click=_clear_report_namespaces,
            )

        if selected_namespaces:
            st.success(
                f"{len(selected_namespaces)} namespace(s) selected: "
                + ", ".join(selected_namespaces)
            )
        else:
            st.info("Select one or more namespaces before generating the report.")

    # ========================================================
    # RESOURCE SELECTION
    # ========================================================

    st.markdown(
        "### 📦 Resources to Include"
    )

    selected_resources = st.multiselect(
        "Select resources",
        [
            "Workloads",
            "Images",
            "ConfigMaps",
            "Secrets",
            "Ingress",
            "Volumes",
            "PVC",
        ],
        default=[
            "Workloads",
            "Images",
            "ConfigMaps",
            "Secrets",
            "Ingress",
            "Volumes",
            "PVC",
        ],
        key="report_resources",
    )

    # ========================================================
    # GENERATE
    # ========================================================

    if st.button(
        "📊 Generate Comparison Report",
        type="primary",
        use_container_width=True,
        key="report_generate",
    ):

        if not selected_resources:

            st.error(
                "Select at least one resource."
            )

            return

        try:

            with st.spinner(
                "Scanning clusters and generating Excel..."
            ):

                if scope == "Selected Namespace" and not selected_namespaces:
                    st.error(
                        "Select at least one namespace."
                    )
                    st.stop()

                report_namespaces = (
                    selected_namespaces
                    if scope == "Selected Namespace"
                    else None
                )

                source_data, destination_data = (
                    collect_both_clusters(
                        source_clients,
                        destination_clients,
                        None,
                        selected_resources,
                        report_namespaces,
                    )
                )

                excel_file = (
                    generate_excel(
                        source_env,
                        destination_env,
                        scope,
                        report_namespaces,
                        source_data,
                        destination_data,
                    )
                )

                st.session_state[
                    "report_excel"
                ] = excel_file.getvalue()

                st.session_state[
                    "report_filename"
                ] = (
                    "cluster_comparison_"
                    f"{source_env}_to_"
                    f"{destination_env}_"
                    f"{('ALL_NAMESPACES' if scope == 'All Namespaces' else 'MULTIPLE_NAMESPACES')}"
                    ".xlsx"
                )

                st.session_state[
                    "report_generated"
                ] = True

            st.success(
                "Comparison report generated successfully."
            )

        except Exception as exc:

            st.error(
                "Failed to generate report."
            )

            st.exception(exc)

    # ========================================================
    # DOWNLOAD
    # ========================================================

    if st.session_state.get(
        "report_generated",
        False,
    ):

        st.download_button(
            "⬇️ Download Excel Report",
            data=st.session_state[
                "report_excel"
            ],
            file_name=st.session_state[
                "report_filename"
            ],
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            use_container_width=True,
            key="report_download",
        )